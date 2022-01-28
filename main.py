import sys
import os
import io

import platform
import sqlite3
import pdfkit

from openpyxl import *
import datetime
from xlsx2html import xlsx2html

from widgets import *
from modules import *

# SET AS GLOBAL WIDGETS
# ///////////////////////////////////////////////////////////////
widgets = None
counter = 0

counter_progress = 0
jumper = 10


# LOGIN WINDOW
# ///////////////////////////////////////////////////////////////
class LoginWindow(QMainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        self.first_name = None
        self.email = None
        self.ui = ui_Login_UI.Ui_MainWindow()
        self.ui.setupUi(self)

        # HIDE SIGN UP RELATED BUTTONS
        # ///////////////////////////////////////////////////////////////
        self.ui.Con_Password_Signup.hide()
        self.ui.First_Name_Signup.hide()
        self.ui.Password_Signup.hide()
        self.ui.Last_Name_Signup.hide()
        self.ui.Email_Signup.hide()
        self.ui.Register_Label.hide()
        self.ui.btn_Register.hide()

        self.ui.Login_Label.setGeometry(QRect(150, 60, 100, 40))
        self.ui.username_input.setGeometry(QRect(100, 140, 190, 40))
        self.ui.Password_input.setGeometry(QRect(100, 200, 190, 40))
        self.ui.Forget_Label.setGeometry(QRect(90, 330, 221, 41))
        self.ui.layoutWidget.move(70, 390)
        self.ui.Login_Input.setGeometry(QRect(100, 280, 190, 40))

        # SET LOGO GIF
        # /////////////////////////////////////////////////////////
        gif_logo = QMovie("images/images/Medtech India pvt ltd.gif")
        gif_logo.setScaledSize(QSize(390, 500))
        self.ui.label_logo.setStyleSheet("background-color: transparent;border-radius: 10px;padding: 5px;")
        self.ui.label_logo.setMovie(gif_logo)

        gif_logo.start()

        # SET UI DEFINITIONS
        # ///////////////////////////////////////////////////////////////
        UIFunctions_Login.uiDefinitions(self)

        self.ui.login_sign_btn.clicked.connect(self.login_sign_btn_clicked)
        self.ui.Login_Input.clicked.connect(self.loginfunction)
        self.ui.btn_Register.clicked.connect(self.signupfunction)

        # SHOW APP
        # ///////////////////////////////////////////////////////////////
        self.show()

    def login_sign_btn_clicked(self):
        if self.ui.login_sign_btn.isChecked():
            self.ui.login_sign_btn.setText("<")
            self.ui.Login_Label.hide()
            self.ui.username_input.hide()
            self.ui.Password_input.hide()
            self.ui.Forget_Label.hide()
            self.ui.Login_Input.hide()

            self.ui.Con_Password_Signup.setGeometry(QRect(30, 280, 355, 40))
            self.ui.First_Name_Signup.setGeometry(QRect(30, 110, 170, 40))
            self.ui.Password_Signup.setGeometry(QRect(30, 220, 355, 40))
            self.ui.Last_Name_Signup.setGeometry(QRect(220, 110, 170, 40))
            self.ui.Email_Signup.setGeometry(QRect(30, 170, 355, 40))
            self.ui.Register_Label.setGeometry(QRect(150, 60, 100, 40))
            self.ui.btn_Register.setGeometry(QRect(100, 350, 190, 40))
            self.ui.layoutWidget.move(70, 410)

            self.ui.Con_Password_Signup.show()
            self.ui.First_Name_Signup.show()
            self.ui.Password_Signup.show()
            self.ui.Last_Name_Signup.show()
            self.ui.Email_Signup.show()
            self.ui.Register_Label.show()
            self.ui.btn_Register.show()
            self.ui.layoutWidget.show()
        else:
            self.ui.Login_Label.show()
            self.ui.username_input.show()
            self.ui.Password_input.show()
            self.ui.Forget_Label.show()
            self.ui.Login_Input.show()
            self.ui.login_sign_btn.setText(">")
            self.ui.Con_Password_Signup.hide()
            self.ui.First_Name_Signup.hide()
            self.ui.Password_Signup.hide()
            self.ui.Last_Name_Signup.hide()
            self.ui.Email_Signup.hide()
            self.ui.Register_Label.hide()
            self.ui.btn_Register.hide()
            self.ui.Login_Label.setGeometry(QRect(150, 60, 100, 40))
            self.ui.username_input.setGeometry(QRect(100, 140, 190, 40))
            self.ui.Password_input.setGeometry(QRect(100, 200, 190, 40))
            self.ui.Forget_Label.setGeometry(QRect(90, 330, 221, 41))
            self.ui.layoutWidget.move(70, 390)
            self.ui.Login_Input.setGeometry(QRect(100, 280, 190, 40))

    def loginfunction(self):
        user = self.ui.username_input.text()
        password = self.ui.Password_input.text()

        if len(user) == 0 or len(password) == 0:
            print("please enter value")
        else:
            conn = sqlite3.connect('user_data.db')
            cur = conn.cursor()
            query = 'SELECT Password FROM DB_USER WHERE Email = \'' + user + '\''
            cur.execute(query)
            try:
                results_password = cur.fetchone()[0]
                if results_password == password:
                    self.close()
                    self.email = user
                    cur.execute(
                        'SELECT FirstName,LastName,UserName,Gender,BirthDate FROM DB_USER WHERE Email = \'' + user + '\'')
                    result = cur.fetchone()
                    self.first_name = result[0]
                    self.last_name = result[1]
                    self.username = result[2]
                    self.Gender = result[3]
                    self.date = result[4]
                    self.date = QDate.fromString(self.date, "yyyy-MM-dd")
                    self.window = MainWindow(self.email, self.first_name, self.last_name, self.username,
                                             self.Gender, self.date)
                    self.window.show()
                else:
                    print("Invalid Password or Username")
            except:
                print("USER not found in database")

    def signupfunction(self):

        first_name = self.ui.First_Name_Signup.text()
        last_name = self.ui.Last_Name_Signup.text()
        email_id = self.ui.Email_Signup.text()
        password = self.ui.Password_Signup.text()
        con_password = self.ui.Con_Password_Signup.text()

        if len(first_name) == 0 or len(last_name) == 0 or len(email_id) == 0 or len(password) == 0 or len(
                con_password) == 0:
            print("please enter value")
        elif password != con_password:
            print("Password does not match")
        else:
            conn = sqlite3.connect('user_data.db')
            cur = conn.cursor()
            user_info = [first_name, last_name, email_id, password]
            cur.execute('INSERT INTO DB_USER (FirstName, LastName, Email, Password) VALUES (?,?,?,?)', user_info)

            conn.commit()
            conn.close()
            print("Signup Successful")


# SPLASH SCREEN
# //////////////////////////////////////////////////////////////////
class SplashWindow(QMainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        self.ui = ui_Splash_screen.Ui_MainWindow()
        self.ui.setupUi(self)

        # UI FUNCTIONS
        # ///////////////////////////////////////////////////////////////
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)

        # DROP SHADOW EFFECT
        # ///////////////////////////////////////////////////////////////
        self.shadow = QGraphicsDropShadowEffect()
        self.shadow.setBlurRadius(20)
        self.shadow.setXOffset(0)
        self.shadow.setYOffset(0)
        self.shadow.setColor(QColor(0, 0, 0, 60))
        self.ui.drop_shadow_frame.setGraphicsEffect(self.shadow)

        # QTIMER FUNCTION
        # ///////////////////////////////////////////////////////////////
        self.timer = QTimer()
        self.timer.timeout.connect(self.progress)

        # TIMER IN MILLISECONDS
        self.timer.start(35)

        # CHANGE DESCRIPTION
        self.ui.label_loading.setText("<strong>Welcome to the</strong> MedTech")
        QtCore.QTimer.singleShot(1500, lambda: self.ui.label_loading.setText("<strong>LOADING</strong> DATABASE"))
        QtCore.QTimer.singleShot(3000, lambda: self.ui.label_loading.setText("<strong>LOADING</strong> USER INTERFACE"))

        # SHOW --------------> MAIN WINDOW
        self.show()

    def progress(self):
        global counter

        # SET VALUE OF PROGRESS BAR

        self.ui.app_progress.setValue(counter)

        counter += 1

        #  CLOSE SPLASH AND START OTHER WINDOW

        if counter > 100:
            # STOP TIMER NOW
            self.timer.stop()

            # CLOSE SPLASH SCREEN
            self.close()

            # OPEN LOGIN WINDOW
            self.login_window = LoginWindow()
            self.login_window.show()


# MAIN APP ------------------------->
# //////////////////////////////////////////////////////////////////
class MainWindow(QMainWindow):
    def __init__(self, email, first_name, last_name, username, Gender, date):
        QMainWindow.__init__(self)
        self.search_query = None
        self.ui = ui_main.Ui_MainWindow()
        self.Gender_Value = Gender
        self.email = email
        self.first_name = first_name
        self.last_name = last_name
        self.username = username
        self.date = date

        self.ui.setupUi(self)

        # UI FUNCTIONS
        # ///////////////////////////////////////////////////////////////
        UIFunctions_main.uiDefinitions_main(self)

        # SHOW PROFILE DETAILS
        self.ui.Email_Input.setText(self.email)
        self.ui.FirstName_Input.setText(self.first_name)
        self.ui.LastName_Input.setText(self.last_name)
        self.ui.UserName_Input.setText(self.username)
        self.ui.Gender_Input.setCurrentText(self.Gender_Value)

        self.ui.BirthDate_Input.setDate(self.date)

        # TOGGLE MENU
        self.ui.toggleButton.clicked.connect(lambda: UIFunctions_main.toggleMenu(self, True))

        # DATE EDIT ENABLE DIALOG
        self.ui.BirthDate_Input.setCursor(Qt.PointingHandCursor)
        self.ui.BirthDate_Input.setCalendarPopup(True)

        # BUTTON CLICK FOR DIFFERENT WINDOWS
        self.ui.btn_profile.clicked.connect(self.btnevent)
        self.ui.btn_labtest.clicked.connect(self.btnevent)
        self.ui.btn_logout.clicked.connect(self.btnevent)
        self.ui.btn_details_save.clicked.connect(self.btnevent)
        self.ui.btn_appointment.clicked.connect(self.btnevent)
        self.ui.btn_search.clicked.connect(self.btnevent)

        # LAB TEST BTN EVENT
        self.ui.btn_covid.clicked.connect(self.btnevent)
        self.ui.btn_cancer.clicked.connect(self.btnevent)
        self.ui.btn_B12_D3.clicked.connect(self.btnevent)
        self.ui.btn_Dengue.clicked.connect(self.btnevent)
        self.ui.btn_ECG.clicked.connect(self.btnevent)
        self.ui.btn_heptatitis.clicked.connect(self.btnevent)
        self.ui.btn_fullbody.clicked.connect(self.btnevent)
        self.ui.btn_hormone.clicked.connect(self.btnevent)
        self.ui.btn_iron.clicked.connect(self.btnevent)
        self.ui.btn_kidney.clicked.connect(self.btnevent)
        self.ui.btn_liver.clicked.connect(self.btnevent)
        self.ui.btn_lungs.clicked.connect(self.btnevent)
        self.ui.btn_stress.clicked.connect(self.btnevent)
        self.ui.btn_thyroid.clicked.connect(self.btnevent)

        self.ui.Gender_Input.activated[int].connect(self.getvalue)
        self.ui.UserName_Input.textChanged.connect(self.uservalue)
        self.ui.BirthDate_Input.dateChanged.connect(self.onDateChanged)

        # STACKED WIDGET HOME PAGE SET
        self.ui.stackedWidget.setCurrentWidget(self.ui.Home_Page)
        self.ui.btn_home.setStyleSheet(UIFunctions_main.selectMenu(self.ui.btn_home.styleSheet()))

        # EXTRA LEFT BOX(CREDIT LABEL)
        def openCloseLeftBox():
            UIFunctions_main.toggleLeftBox(self, True)

        self.ui.toggleLeftBox.clicked.connect(openCloseLeftBox)
        self.ui.extraCloseColumnBtn.clicked.connect(openCloseLeftBox)

        # EXTRA RIGHT BOX( LOGOUT BUTTON)
        def openCloseRightBox():
            UIFunctions_main.toggleRightBox(self, True)

        self.ui.settingsTopBtn.clicked.connect(openCloseRightBox)

        # ===================>*<==============================
        # SHOW WINDOW =======================================>
        self.show()

    # RESIZE EVENTS
    # ///////////////////////////////////////////////////////////////
    def resizeEvent(self, event):
        # Update Size Grips
        UIFunctions_main.resize_grips(self)

    def btnevent(self):
        # SENDER EVENT
        btn = self.sender()
        btn_name = btn.objectName()

        if btn_name == "btn_profile":
            self.ui.stackedWidget.setCurrentWidget(self.ui.Profile_Page)
            UIFunctions_main.resetStyle(self, "btn_profile")
            btn.setStyleSheet(UIFunctions_main.selectMenu(btn.styleSheet()))

        if btn_name == "btn_labtest":
            self.ui.stackedWidget.setCurrentWidget(self.ui.Lab_test_Booking)
            UIFunctions_main.resetStyle(self, "btn_labtest")
            btn.setStyleSheet(UIFunctions_main.selectMenu(btn.styleSheet()))

        if btn_name == "btn_appointment":
            self.ui.stackedWidget.setCurrentWidget(self.ui.Appointment_Page)
            UIFunctions_main.resetStyle(self, "btn_appointment")
            btn.setStyleSheet(UIFunctions_main.selectMenu(btn.styleSheet()))

        if btn_name == "btn_logout":
            self.close()
            self.login_window = LoginWindow()
            self.login_window.show()

        if btn_name == "btn_details_save":
            self.email = self.ui.Email_Input.text()
            conn = sqlite3.connect('user_data.db')
            cur = conn.cursor()
            birth_date = self.date.toString("yyyy-MM-dd")
            user_remain_info = [self.Gender_Value, self.username, birth_date, self.email]
            cur.execute('UPDATE DB_USER SET Gender = ?, UserName= ?,BirthDate=? where Email = ?', user_remain_info)
            conn.commit()
            conn.close()

        if btn_name == "btn_covid":
            self.test_name = "Covid-RT-PCR"
            self.lab_test = LabTestBook(self.first_name, self.last_name, self.test_name)
            self.lab_test.show()

        if btn_name == "btn_cancer":
            self.test_name = "Cancer screening Package"
            self.lab_test = LabTestBook(self.first_name, self.last_name, self.test_name)
            self.lab_test.show()

        if btn_name == "btn_B12_D3":
            self.test_name = "B12/D3 Vitamin Test"
            self.lab_test = LabTestBook(self.first_name, self.last_name, self.test_name)
            self.lab_test.show()

        if btn_name == "btn_Dengue":
            self.test_name = "Dengue Early Detect Test"
            self.lab_test = LabTestBook(self.first_name, self.last_name, self.test_name)
            self.lab_test.show()

        if btn_name == "btn_ECG":
            self.test_name = "ECG Test"
            self.lab_test = LabTestBook(self.first_name, self.last_name, self.test_name)
            self.lab_test.show()

        if btn_name == "btn_fullbody":
            self.test_name = "Full Body Checkup"
            self.lab_test = LabTestBook(self.first_name, self.last_name, self.test_name)
            self.lab_test.show()

        if btn_name == "btn_heptatitis":
            self.test_name = "Heptatitis B AntiBody Test"
            self.lab_test = LabTestBook(self.first_name, self.last_name, self.test_name)
            self.lab_test.show()

        if btn_name == "btn_hormone":
            self.test_name = "Thyroid Hormone Test"
            self.lab_test = LabTestBook(self.first_name, self.last_name, self.test_name)
            self.lab_test.show()

        if btn_name == "btn_iron":
            self.test_name = "Iron - Serum Test"
            self.lab_test = LabTestBook(self.first_name, self.last_name, self.test_name)
            self.lab_test.show()

        if btn_name == "btn_kidney":
            self.test_name = "Glucose/Uric Acid Test/Kidney Function Test"
            self.lab_test = LabTestBook(self.first_name, self.last_name, self.test_name)
            self.lab_test.show()

        if btn_name == "btn_liver":
            self.test_name = "Liver Function Test (LFT)"
            self.lab_test = LabTestBook(self.first_name, self.last_name, self.test_name)
            self.lab_test.show()

        if btn_name == "btn_lungs":
            self.test_name = "Lactate Dehydrogenase (LDH) - Serum Test"
            self.lab_test = LabTestBook(self.first_name, self.last_name, self.test_name)
            self.lab_test.show()

        if btn_name == "btn_thyroid":
            self.test_name = "TSH_T3_T4 Test"
            self.lab_test = LabTestBook(self.first_name, self.last_name, self.test_name)
            self.lab_test.show()

        if btn_name == "btn_stress":
            self.test_name = "Lipoprotein A (LP-A) Test for cholesterol"
            self.lab_test = LabTestBook(self.first_name, self.last_name, self.test_name)
            self.lab_test.show()

        if btn_name == "btn_search":
            self.search_query = self.ui.searchBar.text()
            self.ui.result_widget.clear()
            conn = sqlite3.connect('user_data.db')
            cur = conn.cursor()
            cur.execute("SELECT * FROM Doctor_Data")
            data = cur.fetchall()
            for d in data:
                if self.search_query.startswith(d[2][0:3].lower()) or self.search_query.startswith(d[2][0:3].upper()):
                    item = QListWidgetItem(self.ui.result_widget)
                    self.ui.result_widget.addItem(item)
                    row = MyCustomWidget()
                    item.setSizeHint(row.minimumSizeHint())
                    self.ui.result_widget.setItemWidget(item, row)
                    row.btn_book.clicked.connect(self.handleButtonClicked)
                    row.Doctor_Name_Input.setText(d[1])
                    row.Speciality_Input.setText(d[2])
                    row.Hospital_Name_Input.setText(d[3])
                    row.Experience_Input.setText(str(d[4]) + "years")
                    row.Charge_Input.setText(str(d[5]) + "Rs.")
                    row.Doctor_Pic.setPixmap(QPixmap("images/images/" + d[6]))
                else:
                    pass

    def handleButtonClicked(self):
        widget = self.sender()
        gp = widget.mapToGlobal(QPoint())
        lp = self.ui.result_widget.viewport().mapFromGlobal(gp)
        row = self.ui.result_widget.row(self.ui.result_widget.itemAt(lp))
        self.doctor_name = self.ui.result_widget.itemWidget(self.ui.result_widget.item(row)).Doctor_Name_Input.text()
        self.consultation_charge = self.ui.result_widget.itemWidget(self.ui.result_widget.item(row)).Charge_Input.text()
        self.consultation_charge = int(self.consultation_charge[:-3])
        self.booking_window = BookSlot(self.doctor_name,self.consultation_charge,self.first_name,self.last_name,self.email)
        self.booking_window.show()

    def getvalue(self, value):
        # global Gender_Value
        if value == 1:
            self.Gender_Value = "Male"
        elif value == 2:
            self.Gender_Value = "Female"

    def uservalue(self, value):
        self.username = value

    def onDateChanged(self, newdate):
        self.date = newdate


class MyCustomWidget(QWidget):
    """
    Show Doctor's Information after searching particular doctor
    information from database:
    1. Doctor's Name
    2. Doctor's Specialization
    3. Hospital Name
    4. charge of doctor
    5. Doctor's Photo
    6. Year of experience
    7. there is one button for booking appointment
    """

    def __init__(self, parent=None):
        super(MyCustomWidget, self).__init__(parent)

        self.bg_horizontal = QHBoxLayout()
        self.bg_horizontal.setObjectName(u"bg_horizontal")
        self.bg_horizontal.setContentsMargins(0, 0, 0, 0)
        self.Doctor_Pic = QLabel()
        self.Doctor_Pic.setObjectName(u"Doctor_Pic")
        self.Doctor_Pic.setMinimumSize(QSize(150, 150))
        self.Doctor_Pic.setMaximumSize(QSize(150, 150))

        self.bg_horizontal.addWidget(self.Doctor_Pic)

        self.bg_part_1 = QVBoxLayout()
        self.bg_part_1.setObjectName(u"bg_part_1")
        self.Doctor_Name_Layout = QHBoxLayout()
        self.Doctor_Name_Layout.setObjectName(u"Doctor_Name_Layout")
        self.Doctor_Name_label = QLabel()
        self.Doctor_Name_label.setObjectName(u"Doctor_Name_label")
        self.Doctor_Name_label.setMinimumSize(QSize(140, 45))
        self.Doctor_Name_label.setMaximumSize(QSize(140, 45))
        font = QFont()
        font.setPointSize(14)
        font.setBold(True)
        self.Doctor_Name_label.setFont(font)

        self.Doctor_Name_Layout.addWidget(self.Doctor_Name_label)

        self.Doctor_Name_Input = QLineEdit()
        self.Doctor_Name_Input.setObjectName(u"Doctor_Name_Input")
        self.Doctor_Name_Input.setMinimumSize(QSize(180, 45))
        self.Doctor_Name_Input.setMaximumSize(QSize(180, 45))

        self.Doctor_Name_Layout.addWidget(self.Doctor_Name_Input)

        self.bg_part_1.addLayout(self.Doctor_Name_Layout)

        self.Speciality_Layout = QHBoxLayout()
        self.Speciality_Layout.setObjectName(u"Speciality_Layout")
        self.Speciality_Label = QLabel()
        self.Speciality_Label.setObjectName(u"Speciality_Label")
        self.Speciality_Label.setMinimumSize(QSize(140, 45))
        self.Speciality_Label.setMaximumSize(QSize(140, 45))
        self.Speciality_Label.setFont(font)

        self.Speciality_Layout.addWidget(self.Speciality_Label)

        self.Speciality_Input = QLineEdit()
        self.Speciality_Input.setObjectName(u"Speciality_Input")
        self.Speciality_Input.setMinimumSize(QSize(180, 45))
        self.Speciality_Input.setMaximumSize(QSize(180, 45))

        self.Speciality_Layout.addWidget(self.Speciality_Input)

        self.bg_part_1.addLayout(self.Speciality_Layout)

        self.Hospital_Name_Layout = QHBoxLayout()
        self.Hospital_Name_Layout.setObjectName(u"Hospital_Name_Layout")
        self.Hospital_Name_Label = QLabel()
        self.Hospital_Name_Label.setObjectName(u"Hospital_Name_Label")
        self.Hospital_Name_Label.setMinimumSize(QSize(140, 45))
        self.Hospital_Name_Label.setMaximumSize(QSize(140, 45))
        self.Hospital_Name_Label.setFont(font)

        self.Hospital_Name_Layout.addWidget(self.Hospital_Name_Label)

        self.Hospital_Name_Input = QLineEdit()
        self.Hospital_Name_Input.setObjectName(u"Hospital_Name_Input")
        self.Hospital_Name_Input.setMinimumSize(QSize(180, 45))
        self.Hospital_Name_Input.setMaximumSize(QSize(180, 45))

        self.Hospital_Name_Layout.addWidget(self.Hospital_Name_Input)

        self.bg_part_1.addLayout(self.Hospital_Name_Layout)

        self.bg_horizontal.addLayout(self.bg_part_1)

        self.bg_part_2 = QVBoxLayout()
        self.bg_part_2.setObjectName(u"bg_part_2")
        self.YOE_Layout = QHBoxLayout()
        self.YOE_Layout.setObjectName(u"YOE_Layout")
        self.Experience_Label = QLabel()
        self.Experience_Label.setObjectName(u"Experience_Label")
        self.Experience_Label.setMinimumSize(QSize(190, 45))
        self.Experience_Label.setMaximumSize(QSize(190, 45))
        self.Experience_Label.setFont(font)

        self.YOE_Layout.addWidget(self.Experience_Label)

        self.Experience_Input = QLineEdit()
        self.Experience_Input.setObjectName(u"Experience_Input")
        self.Experience_Input.setMinimumSize(QSize(190, 45))
        self.Experience_Input.setMaximumSize(QSize(190, 45))

        self.YOE_Layout.addWidget(self.Experience_Input)

        self.bg_part_2.addLayout(self.YOE_Layout)

        self.charge_Layout = QHBoxLayout()
        self.charge_Layout.setObjectName(u"charge_Layout")
        self.Charge_Label = QLabel()
        self.Charge_Label.setObjectName(u"Charge_Label")
        self.Charge_Label.setMinimumSize(QSize(190, 45))
        self.Charge_Label.setMaximumSize(QSize(190, 45))
        self.Charge_Label.setFont(font)

        self.charge_Layout.addWidget(self.Charge_Label)

        self.Charge_Input = QLineEdit()
        self.Charge_Input.setObjectName(u"Charge_Input")
        self.Charge_Input.setMinimumSize(QSize(190, 45))
        self.Charge_Input.setMaximumSize(QSize(190, 45))

        self.charge_Layout.addWidget(self.Charge_Input)

        self.bg_part_2.addLayout(self.charge_Layout)

        self.btn_layout = QHBoxLayout()
        self.btn_layout.setObjectName(u"btn_layout")
        self.btn_space = QSpacerItem(230, 45, QSizePolicy.Expanding, QSizePolicy.Minimum)

        self.btn_layout.addItem(self.btn_space)

        self.btn_book = QPushButton()
        self.btn_book.setObjectName(u"btn_book")
        self.btn_book.setMinimumSize(QSize(150, 45))
        self.btn_book.setMaximumSize(QSize(150, 45))

        self.btn_layout.addWidget(self.btn_book)

        self.bg_part_2.addLayout(self.btn_layout)

        self.bg_horizontal.addLayout(self.bg_part_2)

        self.setLayout(self.bg_horizontal)

        # SET LABEL TEXT
        self.Doctor_Pic.setText("")
        self.Doctor_Name_label.setText(u"Doctor Name :")
        self.Speciality_Label.setText(u"Speciality :")
        self.Hospital_Name_Label.setText(u"Hospital Name:")
        self.Experience_Label.setText(u"Total Experience :")
        self.Charge_Label.setText(u"consultation Charge :")
        self.btn_book.setText(u"Book")


# BOOK SCREEN WINDOW
# ######################################################
class BookSlot(QMainWindow):
    def __init__(self,doctor_name,consultation_charge,first_name,last_name,email):
        QMainWindow.__init__(self)
        self.first_name = first_name
        self.email = email
        self.last_name = last_name
        self.date = None
        self.slot_name = 0
        self.token_count = 0
        self.doctor_name = doctor_name
        self.consultation_charge = consultation_charge
        self.ui = ui_Booking.Ui_MainWindow()
        self.ui.setupUi(self)

        self.setWindowFlags(Qt.FramelessWindowHint)        # STANDARD TITLE BAR REMOVED
        self.setAttribute(Qt.WA_TranslucentBackground)     # TRANSPARENT BACKGROUND

        # MINIMIZE
        self.ui.minimizeAppBtn.clicked.connect(lambda: self.showMinimized())

        # CLOSE APPLICATION
        self.ui.closeAppBtn.clicked.connect(lambda: self.close())

        # DATE EDIT ENABLE DIALOG
        self.ui.Date_Selector.setCursor(Qt.PointingHandCursor)
        self.ui.Date_Selector.setCalendarPopup(True)

        # Date EDIT ENABLE DIALOG
        self.ui.Date_Selector.dateChanged.connect(self.onDateChanged)

        self.ui.Doc_Name.setText("Avilable slot for " + self.doctor_name)

        self.ui.btn_slot_01.clicked.connect(self.datechange)
        self.ui.btn_slot_02.clicked.connect(self.datechange)
        self.ui.btn_slot_03.clicked.connect(self.datechange)
        self.ui.btn_slot_04.clicked.connect(self.datechange)
        self.ui.btn_slot_05.clicked.connect(self.datechange)
        self.ui.btn_slot_06.clicked.connect(self.datechange)
        self.ui.btn_slot_07.clicked.connect(self.datechange)
        self.ui.btn_slot_08.clicked.connect(self.datechange)
        self.ui.btn_slot_09.clicked.connect(self.datechange)
        self.ui.btn_slot_10.clicked.connect(self.datechange)
        self.ui.btn_slot_11.clicked.connect(self.datechange)
        self.ui.btn_slot_12.clicked.connect(self.datechange)
        # SHOW APP
        # ///////////////////////////////////////////////////////////////
        self.show()

    def onDateChanged(self,newdate):
        self.date = newdate
        self.date.toString("yyyy.MM.dd")
        select_date = self.date.toString("yyyy-MM-dd")
        doctor_name = ["Dr. Asutosh Sharma","Dr. Parul Agrawal","Dr. Rita Patel","Dr. Mayank","Dr. Shikha Manek","Dr. Vishal Modi","Dr. Smriti Pandya","Dr. Anushka Patil","Dr. Viren patel","Dr. Piyush Sharma"]
        doctor_database_name = ["Dr_Asutosh_Sharma_Slots","Dr_Parul_Agrawal_Slots","Dr_Rita_Patel_Slots","Dr_Mayank_Slots","Dr_Shikha_Manek_Slots","Dr_Vishal_Modi_Slots","Dr_Smriti_Pandya_Slots","Dr_Anushka_Patil_Slots","Dr_Viren_Patel_Slots","Dr_Piyush_Sharma_Slots"]
        btnname = ["btn_slot_01", "btn_slot_02", "btn_slot_03", "btn_slot_04", "btn_slot_05", "btn_slot_06",
                   "btn_slot_07", "btn_slot_08", "btn_slot_09", "btn_slot_10", "btn_slot_11", "btn_slot_12"]

        conn = sqlite3.connect('user_data.db')
        cur = conn.cursor()
        for i in range(len(doctor_name)):
            if doctor_name[i] == self.doctor_name:
                query = 'SELECT * FROM ' + doctor_database_name[i] + ' WHERE Day = \'' + select_date + '\''
                cur.execute(query)
                results = cur.fetchone()
                for j in range(2,len(results)):
                    if results[j] == 10:
                        self.ui.__dict__[btnname[j - 2]].setEnabled(False)
                        self.ui.__dict__[btnname[j - 2]].setStyleSheet(
                            "background-color: rgba(98, 114, 164, 120);border-color: rgba(98, 114, 164, 120)")
                    else:
                        self.ui.__dict__[btnname[j - 2]].setEnabled(True)
                        self.ui.__dict__[btnname[j - 2]].setStyleSheet(
                            "background-color: rgba(98, 114, 164, 255);border-color: rgba(98, 114, 164, 255)")

    def datechange(self):
        btn = self.sender()
        btn_name = btn.objectName()
        select_date = self.date.toString("yyyy-MM-dd")
        conn = sqlite3.connect('user_data.db')
        cur = conn.cursor()
        doctor_name = ["Dr. Asutosh Sharma","Dr. Parul Agrawal","Dr. Rita Patel","Dr. Mayank","Dr. Shikha Manek","Dr. Vishal Modi","Dr. Smriti Pandya","Dr. Anushka Patil","Dr. Viren patel","Dr. Piyush Sharma"]
        doctor_database_name = ["Dr_Asutosh_Sharma_Slots","Dr_Parul_Agrawal_Slots","Dr_Rita_Patel_Slots","Dr_Mayank_Slots","Dr_Shikha_Manek_Slots","Dr_Vishal_Modi_Slots","Dr_Smriti_Pandya_Slots","Dr_Anushka_Patil_Slots","Dr_Viren_Patel_Slots","Dr_Piyush_Sharma_Slots"]
        btnname = ["btn_slot_01", "btn_slot_02", "btn_slot_03", "btn_slot_04", "btn_slot_05", "btn_slot_06", "btn_slot_07", "btn_slot_08", "btn_slot_09", "btn_slot_10", "btn_slot_11", "btn_slot_12"]
        slot_name =[0,1,2,3,4,5,6,7,8,9,10,11]
        for i in range(len(doctor_name)):
            if doctor_name[i] == self.doctor_name:
                query = 'SELECT * FROM ' + doctor_database_name[i] + ' WHERE Day = \'' + select_date + '\''
                cur.execute(query)
                results = cur.fetchone()
                for j in range(len(btnname)):
                    if btnname[j] == btn_name:
                        if results[j+2] <=10 :
                            self.slot_name = slot_name[j]
                            self.token_count = results[j+2]
                self.close()
                self.Book_Info = BookInfo(self.doctor_name,select_date,self.consultation_charge,self.first_name,self.last_name,self.email,self.slot_name,self.token_count)
                self.Book_Info.show()

# DOCTOR APPOINTMENT BOOKING INFORMATION
# ##########################################################
class BookInfo(QMainWindow):
    def __init__(self,doctor_name,select_date,consultation_charge,first_name,last_name,email,slot_name,token_count):
        QMainWindow.__init__(self)
        self.doctor_name = doctor_name
        self.select_date = select_date
        self.consultation_charge = consultation_charge
        self.first_name = first_name
        self.last_name = last_name
        self.email = email
        self.slot_name = slot_name
        self.token_count = token_count
        self.ui = ui_Booking_Info.Ui_MainWindow()
        self.ui.setupUi(self)

        # STANDARD TITLE BAR
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)

        # MINIMIZE
        self.ui.minimizeAppBtn.clicked.connect(lambda: self.showMinimized())

        # CLOSE APPLICATION
        self.ui.closeAppBtn.clicked.connect(lambda: self.close())

        # SET DETAILS
        self.ui.name_input.setText(self.first_name)
        self.ui.lastname_input.setText(self.last_name)
        self.ui.charge_Input.setText(str(self.consultation_charge) + " Rs")

        self.ui.label_title.setText(
            "Dear " + first_name + "  " + last_name + ", You are Booking appointment on "+self.select_date+" With "+self.doctor_name)

        # BTN CLICK EVENT FOR BOOKING TEST
        self.ui.btn_payLater.clicked.connect(self.btnevent)

        # SHOW WINDOW
        self.show()

    def btnevent(self):
        # SENDER EVENT
        btn = self.sender()
        btn_name = btn.objectName()
        conn = sqlite3.connect('user_data.db')
        cur = conn.cursor()

        if btn_name == "btn_payLater":
            self.paymentstatus = "Pending"
            self.phone_numer = self.ui.mobile_input.text()
            self.Address = self.ui.address_input.text()
            self.city = self.ui.city_input.text()
            doctor_name = ["Dr. Asutosh Sharma","Dr. Parul Agrawal","Dr. Rita Patel","Dr. Mayank","Dr. Shikha Manek","Dr. Vishal Modi","Dr. Smriti Pandya","Dr. Anushka Patil","Dr. Viren patel","Dr. Piyush Sharma"]
            doctor_database_name = ["Dr_Asutosh_Sharma_Slots","Dr_Parul_Agrawal_Slots","Dr_Rita_Patel_Slots","Dr_Mayank_Slots","Dr_Shikha_Manek_Slots","Dr_Vishal_Modi_Slots","Dr_Smriti_Pandya_Slots","Dr_Anushka_Patil_Slots","Dr_Viren_Patel_Slots","Dr_Piyush_Sharma_Slots"]
            slot_name = ["slot_1", "slot_2", "slot_3", "slot_4", "slot_5", "slot_6", "slot_7", "slot_8", "slot_9", "slot_10", "slot_11", "slot_12"]

            if  len(self.phone_numer) == 0 or len(self.Address) == 0 or len(self.city) == 0:
                print("please enter value")
            else:
                for i in range(len(doctor_name)):
                    if doctor_name[i] == self.doctor_name:
                        query = 'SELECT * FROM ' + doctor_database_name[i] + ' WHERE Day = \'' + self.select_date + '\''
                        cur.execute(query)
                        results = cur.fetchone()
                        if results[self.slot_name+2] <= 10:
                            query = 'UPDATE ' + doctor_database_name[i] + ' SET ' + slot_name[self.slot_name] + ' = ' + str(self.token_count)+ ' + 1 WHERE Day = \'' + self.select_date + '\''
                            cur.execute(query)
                            conn.commit()

                        appointment_data = [(self.slot_name+1)*100+self.token_count,self.first_name,self.last_name,self.email,self.phone_numer,self.Address,
                                     self.city,doctor_name[i],self.consultation_charge,self.paymentstatus,self.select_date,self.slot_name+1]
                        cur1 = conn.cursor()
                        cur1.execute('INSERT INTO Doctor_Appointment (TokenNo, FirstName, LastName, Email, MobileNo,Address,City,DoctorName,ConsultationCharge,PaymentStatus,AppointmentDate,AppointmentSlot) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)',appointment_data)
                        conn.commit()
                        conn.close()

                        # INVOICE EXCEL FILE
                        wb = load_workbook("Invoice/Invoice_Format/Invoice_Format_Doctor_Appointment.xlsx")
                        ws = wb["Invoice"]
                        ws["D9"] = self.first_name + " " + self.last_name
                        ws["D10"] = self.Address
                        ws["D11"] = self.city
                        ws["D12"] = self.phone_numer
                        ws["B5"] = "INV0" + str((self.slot_name + 1) * 100 + self.token_count)
                        ws["E5"] = (self.slot_name + 1) * 100 + self.token_count
                        ws["E9"] = self.slot_name + 1
                        ws["F5"] = self.select_date
                        ws["F9"] = doctor_name[i]
                        ws["G19"] = self.consultation_charge
                        ws["G22"] = self.consultation_charge
                        ws["G24"] = (self.consultation_charge * 9) / 100
                        ws["G25"] = (self.consultation_charge * 9) / 100
                        ws["G26"] = (self.consultation_charge * 18) / 100
                        ws["E31"] = self.consultation_charge + (self.consultation_charge * 18) / 100
                        ws["B19"] = " consultation charge for " + doctor_name[i]
                        ws["D5"] = datetime.date.today()
                        self.file_name = "INV0" + str((self.slot_name + 1) * 100 + self.token_count)
                        wb.save("Invoice/Invoice_Lab_Test/Invoice_Excel/" + self.file_name + ".xlsx")
                        self.close()
                        self.process_bar = ProcessBar(self.file_name)
                        self.process_bar.show()


# LAB TEST BOOKING CLASS
# ######################################################
class LabTestBook(QMainWindow):
    """
    THIS CLASS IS USED TO CREATE LAB TEST BOOKING WINDOW
    :VARIABLES:
    first_name: FIRST NAME OF PATIENT
    last_name: LAST NAME OF PATIENT
    test_name: LAB TEST NAME
    phone_number : PHONE NUMBER OF PATIENT
    address: ADDRESS OF PATIENT
    city: CITY OF PATIENT
    PaymentStatus: PAYMENT STATUS OF LAB TEST THAT PATIENT BOOKED WITH SYSTEM
    charge : CHARGE OF LAB TEST
    Invoice_No: INVOICE NUMBER OF LAB TEST
    """

    def __init__(self, first_name, last_name, test_name):
        QMainWindow.__init__(self)
        self.first_name = first_name
        self.last_name = last_name
        self.test_name = test_name
        self.phone_numer = None
        self.city = None
        self.Address = None
        self.charge = 0
        self.PaymentStatus = "Not Done"
        self.Invoice_No = None
        self.ui = ui_LabTestBook.Ui_MainWindow()
        self.ui.setupUi(self)

        # STANDARD TITLE BAR
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)

        # MINIMIZE
        self.ui.minimizeAppBtn.clicked.connect(lambda: self.showMinimized())

        # CLOSE APPLICATION
        self.ui.closeAppBtn.clicked.connect(lambda: self.close())

        # SET DETAILS
        self.ui.name_input.setText(self.first_name)
        self.ui.lastname_input.setText(self.last_name)


        self.ui.label_title.setText(
            "Dear " + first_name + "  " + last_name + ", You are Booking " + test_name + " With us.")

        # BTN CLICK EVENT FOR BOOKING TEST
        self.ui.btn_payLater.clicked.connect(self.btnevent)

        # SHOW APP
        # ///////////////////////////////////////////////////////////////
        self.show()

    def btnevent(self):
        # SENDER EVENT
        btn = self.sender()
        btn_name = btn.objectName()

        if btn_name == "btn_payLater":
            self.phone_numer = self.ui.mobile_input.text()
            self.Address = self.ui.address_input.text()
            self.city = self.ui.city_input.text()
            if len(self.first_name) == 0 or len(self.last_name) == 0 or len(self.phone_numer) == 0 or len(
                    self.Address) == 0 or len(self.city) == 0:
                print("please enter value")
            else:
                if self.test_name == "Covid-RT-PCR":
                    self.charge = 579

                if self.test_name ==  "B12/D3 Vitamin Test":
                    self.charge = 999

                if self.test_name == "Cancer screening Package":
                    self.charge = 4799

                if self.test_name == "Dengue Early Detect Test":
                    self.charge = 899

                if self.test_name == "ECG Test":
                    self.charge = 159

                if self.test_name == "Heptatitis B AntiBody Test":
                    self.charge = 1199

                if self.test_name == "Full Body Checkup":
                    self.charge = 8999

                if self.test_name == "Thyroid Hormone TestIron - Serum":
                    self.charge = 599

                if self.test_name == "Iron - Serum Test":
                    self.charge = 299

                if self.test_name == "Liver Function Test (LFT)":
                    self.charge = 699

                if self.test_name == "Glucose/Uric Acid Test/Kidney Function Test":
                    self.charge = 1099

                if self.test_name == "Lactate Dehydrogenase (LDH) - Serum Test":
                    self.charge = 349

                if self.test_name == "TSH_T3_T4 Test":
                    self.charge = 599

                if self.test_name == "Lipoprotein A (LP-A) Test for cholesterol":
                    self.charge = 1100

                conn = sqlite3.connect('user_data.db')
                cur = conn.cursor()
                user_info = [self.first_name, self.last_name, self.phone_numer, self.city, self.Address, self.charge,
                             self.PaymentStatus]
                cur.execute(
                    'INSERT INTO LabTest_Record (FirstName, LastName, MobileNo, City,Address,Charge,PaymentStatus) VALUES (?,?,?,?,?,?,?)',
                    user_info)

                cur1 = conn.cursor()
                cur1.execute(
                    'SELECT InvoiceNo FROM LabTest_Record WHERE MobileNo = \'' + self.phone_numer + '\'')

                invoiceNo = cur1.fetchone()[0]

                conn.commit()
                conn.close()
                wb = load_workbook("Invoice/Invoice_Format/Invoice_Format_Lab_Test.xlsx")
                ws = wb["Invoice"]
                ws["D9"] = self.first_name + " " + self.last_name
                ws["D10"] = self.Address
                ws["D11"] = self.city
                ws["D12"] = self.phone_numer
                ws["B19"] = self.test_name
                ws["G19"] = self.charge
                ws["G22"] = self.charge
                ws["G24"] = (self.charge * 9) / 100
                ws["G25"] = (self.charge * 9) / 100
                ws["G26"] = (self.charge * 18) / 100
                ws["E31"] = self.charge + (self.charge * 18) / 100
                ws["D5"] = datetime.date.today()
                ws["B5"] = "INV" + "0000" + str(invoiceNo)
                self.file_name = "INV" + "0000" + str(invoiceNo)
                wb.save("Invoice/Invoice_Lab_Test/Invoice_Excel/"+self.file_name+".xlsx")

                self.close()
                self.process_bar = ProcessBar(self.file_name)
                self.process_bar.show()


class ProcessBar(QMainWindow):
    def __init__(self, file_name):
        QMainWindow.__init__(self)
        self.file_name = file_name
        self.ui = ui_process_bar.Ui_SplashScreen()
        self.ui.setupUi(self)

        # ==> SET INITIAL PROGRESS BAR TO (0) ZERO
        self.progressBarValue(0)

        # ==> REMOVE STANDARD TITLE BAR
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)  # Remove title bar
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)  # Set background to transparent

        # ==> APPLY DROP SHADOW EFFECT
        self.shadow = QGraphicsDropShadowEffect(self)
        self.shadow.setBlurRadius(20)
        self.shadow.setXOffset(0)
        self.shadow.setYOffset(0)
        self.shadow.setColor(QColor(0, 0, 0, 120))
        self.ui.circularBg.setGraphicsEffect(self.shadow)

        # QTIMER ==> START
        self.timer = QtCore.QTimer()
        self.timer.timeout.connect(self.progress)
        # TIMER IN MILLISECONDS
        self.timer.start(15)

        # SHOW WINDOW
        self.show()

    # DEF TO LOADING
    ########################################################################
    def progress(self):
        global counter_progress
        global jumper
        value = counter_progress

        # HTML TEXT PERCENTAGE
        htmlText = """<p><span style=" font-size:68pt;">{VALUE}</span><span style=" font-size:58pt; vertical-align:super;">%</span></p>"""

        # REPLACE VALUE
        newHtml = htmlText.replace("{VALUE}", str(jumper))

        if value > jumper:
            # APPLY NEW PERCENTAGE TEXT
            self.ui.labelPercentage.setText(newHtml)
            jumper += 10

        # SET VALUE TO PROGRESS BAR
        # fix max value error if > than 100
        if value >= 100:
            value = 1.000
        self.progressBarValue(value)

        # CLOSE SPLASH SCREEN AND OPEN APP
        if counter_progress > 100:
            # STOP TIMER
            self.timer.stop()

            xlsx2html("Invoice/Invoice_Lab_Test/Invoice_Excel/" + self.file_name+".xlsx", "Invoice/Invoice_Lab_Test/Invoice_Html/"+self.file_name+".html")
            path_wkhtmltopdf = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'
            config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)

            pdfkit.from_file("Invoice/Invoice_Lab_Test/Invoice_Html/"+self.file_name+".html", "Invoice/Invoice_Lab_Test/Invoice_Pdf/"+self.file_name+".pdf", configuration= config)

            self.close()
            self.view_pdf_btn = ViewPdf_btn(self.file_name)
            self.view_pdf_btn.show()

        # INCREASE COUNTER
        counter_progress += 0.5

    # DEF PROGRESS BAR VALUE
    ########################################################################
    def progressBarValue(self, value):

        # PROGRESSBAR STYLESHEET BASE
        styleSheet = """
        QFrame{
            border-radius: 150px;
            background-color: qconicalgradient(cx:0.5, cy:0.5, angle:90, stop:{STOP_1} rgba(255, 0, 127, 0), stop:{STOP_2} rgba(85, 170, 255, 255));
        }
        """

        # GET PROGRESS BAR VALUE, CONVERT TO FLOAT AND INVERT VALUES
        # stop works of 1.000 to 0.000
        progress = (100 - value) / 100.0

        # GET NEW VALUES
        stop_1 = str(progress - 0.001)
        stop_2 = str(progress)

        # SET VALUES TO NEW STYLESHEET
        newStylesheet = styleSheet.replace("{STOP_1}", stop_1).replace("{STOP_2}", stop_2)

        # APPLY STYLESHEET WITH NEW VALUES
        self.ui.circularProgress.setStyleSheet(newStylesheet)

# VIEW PDF BUTTON AND BOOK ANOTHER TEST BUTTON WINDOW
# #############################################################################
class ViewPdf_btn(QMainWindow):
    def __init__(self, file_name):
        QMainWindow.__init__(self)
        self.file_name = file_name
        self.ui = ui_view_pdf_btn.Ui_MainWindow()
        self.ui.setupUi(self)

        # SET UI DEFINITION TO WINDOW
        UIFunctions_pdf_View_btn.uiDefinitions_pdf_View_btn(self)

        # BTN CLICK EVENT
        self.ui.btn_view.clicked.connect(self.btnevent)
        self.ui.btn_cancel.clicked.connect(self.btnevent)
        self.ui.btn_book_other.clicked.connect(self.btnevent)

        # ==> SHOW WINDOW
        self.show()

    def btnevent(self):
        sender = self.sender()
        btn_name = sender.objectName()

        if btn_name == "btn_view":
            self.close()
            self.view_pdf = ViewPdf(self.file_name)
            self.view_pdf.show()

        if btn_name == "btn_cancel":
            self.close()

        if btn_name == "btn_book_other":
            self.close()

# VIEW PDF WINDOW
# #############################################################################
class ViewPdf(QMainWindow):
    def __init__(self, file_name):
        QMainWindow.__init__(self)
        self.file_name = file_name
        self.ui = ui_Invoice_Viewer.Ui_MainWindow()
        self.ui.setupUi(self)

        # SET UI DEFINITION TO WINDOW
        UIFunctions_pdf_View.uiDefinitions_pdf_View(self)

        # SHOW PDF
        self.pdf_layout = QVBoxLayout(self.ui.content_frame)
        self.pdf_layout.setContentsMargins(0, 0, 0, 0)
        self.pdf_viewer = QWebEngineView(self.ui.content_frame)
        self.pdf_layout.addWidget(self.pdf_viewer)
        self.pdf_viewer.settings().setAttribute(QWebEngineSettings.PluginsEnabled, True)
        self.pdf_viewer.load(QUrl.fromLocalFile("D:/MY_PROJECTS/MedTech/Invoice/Invoice_Lab_Test/Invoice_Pdf/"+self.file_name+".pdf"))

        # ==> SHOW WINDOW
        self.show()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = SplashWindow()
    window.show()
    sys.exit(app.exec())
